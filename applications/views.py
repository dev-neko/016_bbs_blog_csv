import csv
import io
import json
import zipfile
from pprint import pprint
import openpyxl as openpyxl
# import pandas as pd
from io import BytesIO
import numpy as np
# Django
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.shortcuts import redirect
from django.views.generic import (
	FormView,
	ListView,
	UpdateView,
	TemplateView,
	CreateView,
)
from django.http import HttpResponse,JsonResponse
# models
from .models import CompanyModel
# forms
from .forms import FileUploadForm
# scripts
from scripts.main import ScrapeToCsv



# ExcelファイルをアップロードしてWebページに表示する
class FileUploadView(LoginRequiredMixin,FormView):
	template_name='applications/index.html'
	form_class=FileUploadForm
	success_url=reverse_lazy('index')
	# context_object_name='ctx'
	# model=CompanyModel

	def form_valid(self,form):
		# フォームから受け取ったデータをデータフレームへ変換
		# ファイル形式に応じた読み込み
		# file=io.TextIOWrapper(form.cleaned_data['file'],encoding='cp932')
		# print(file.encoding)

		# def get_context_data(self,**kwargs):
		# 	ctx=super().get_context_data(**kwargs)
		# 	# excel=self.request.FILES['file']
		# 	return ctx,self.header_cells

		# df=pd.read_excel(file,dtype=str,index_col=0)
		# print(df)

		# with open(file,encoding='cp932') as file:
		# 	data = file.read()

		# wb = xlrd.open_workbook(file, encoding_override='cp932')
		# df = pd.read_excel(wb)
		# print(df)

		# wb=xlwings.Book(df)
		# data_sheet=wb.sheets['Sheet1']
		# test_data=data_sheet.range(1,11).expand('down').value
		# print(test_data)

		# if filetype == 'csv':
		# 	df = pd.read_csv(file, dtype=str)
		# elif filetype == 'xlsx':
		# 	df = pd.read_excel(file, dtype=str, index_col=0)

		excel=self.request.FILES['file']
		# print(excel.read())

		wb=openpyxl.load_workbook(excel)
		ws=wb['Sheet1']
		# print(ws.cell(4,11).value)

		# １行目（列名のセル）
		header_cells = ws[3]
		# print(header_cells[10].value)

		# ２行目以降（データ）
		company_list = []
		for row in ws.iter_rows(min_row=4):
			row_dic = {}
			# セルの値を「key-value」で登録
			for k, v in zip(header_cells, row):
				if k.value!=None:
					row_dic[k.value] = v.value
			company_list.append(row_dic)

		# pprint(student_list)

		# DBに保存
		CompanyModel.objects.update_or_create(defaults={'information':company_list})

		return redirect('app_urls:index')


	# def get_object(self,queryset=None):
	# 	ctx['company_list']=CompanyModel.objects.all()
	# 	print(ctx)
	# 	return ctx

	def get_context_data(self,**kwargs):
		ctx=super().get_context_data(**kwargs)
		company_model_obj=CompanyModel.objects.all()
		try:
			ctx['company_list']=eval(company_model_obj[0].information)
			ctx['title_list']=['企業名', 'HP', '業界', '住所', '従業員数', '設立年月', '上場区分', '総合評価', 'A.事業戦略', 'B.経営手腕', 'C.職場環境', 'D.仕事の意義', 'E.教育・成長', 'F.給与・処遇', 'G.生活しやすさ']
		except:
			pass
		return ctx



class Sample01(TemplateView):
	template_name='applications/falcon-v3.16.0/public/sample01.html'


class v1(LoginRequiredMixin,TemplateView):
	template_name='applications/index_v1.html'
	# template_name='applications/Falcon_nav_test.html'

class SearchFormView(
	# LoginRequiredMixin,
	# FormView,
	TemplateView,):
	template_name='applications/index_v1.html'
	# success_url=reverse_lazy('SearchFormView')
	# form_class = FileUploadForm
	# model=CompanyModel

	def post(self,request):
		req_data=request.POST
		# print(req_data)
		req_textarea_siteurl=req_data.get('textarea_siteurl').split('\r\n')
		req_textarea_excludeword=req_data.get('textarea_excludeword').split(' ')
		# print(req_textarea_siteurl)
		# print(req_textarea_excludeword)

		stc=ScrapeToCsv(req_textarea_excludeword,req_textarea_siteurl)
		csv_res_list_list=stc.main()
		# pprint(csv_res_list_list,sort_dicts=False)

		memory_file=BytesIO()
		zip_file=zipfile.ZipFile(memory_file,'w')

		for count,csv_res_list in enumerate(csv_res_list_list):
			# csv作成、zipにまとめる
			csv_file=HttpResponse(content_type='text/csv')
			writer=csv.writer(csv_file)
			# NumPyで1列多行の2次元リストに変換しないと縦に書き込めない
			np_csv_res_list=np.array(csv_res_list).reshape(-1,1).tolist()
			writer.writerows(np_csv_res_list)
			# print(csv_file.getvalue())
			# zipにまとめる
			zip_file.writestr(f'{count}.csv',csv_file.getvalue())
			csv_file.close()

		# zipファイルの内容をreponseに設定
		zip_file.close()  #ここでcloseしないとエラー発生して解凍できない
		response=HttpResponse(memory_file.getvalue(),content_type='application/zip')
		# videoidを,で区切ってファイル名にする
		# response['Content-Disposition']=f'attachment; filename="{",".join(request.POST.getlist("videoids"))}.zip"'
		response['Content-Disposition']=f'attachment; filename="test.zip"'

		return response
		# return redirect('app_urls:v1')


def axios_form(request):
	if request.method=='POST':
		req_data=json.loads(request.body)
		# print(req_data)
		req_textarea_siteurl=req_data.get('textarea_siteurl').split('\n')
		req_textarea_excludeword=req_data.get('textarea_excludeword').split(' ')
		# print(req_textarea_siteurl)
		# print(req_textarea_excludeword)

		stc=ScrapeToCsv(req_textarea_excludeword,req_textarea_siteurl)
		csv_res_list_list=stc.main()
		pprint(csv_res_list_list,sort_dicts=False)

		memory_file=BytesIO()
		zip_file=zipfile.ZipFile(memory_file,'w')

		for count,csv_res_list in enumerate(csv_res_list_list):
			# csv作成、zipにまとめる
			csv_file=HttpResponse(content_type='text/csv')
			writer=csv.writer(csv_file)
			# NumPyで1列多行の2次元リストに変換しないと縦に書き込めない
			np_csv_res_list=np.array(csv_res_list).reshape(-1,1).tolist()
			writer.writerows(np_csv_res_list)
			# print(csv_file.getvalue())
			# zipにまとめる
			zip_file.writestr(f'{count}.csv',csv_file.getvalue())
			csv_file.close()

		# zipファイルの内容をreponseに設定
		zip_file.close()  #ここでcloseしないとエラー発生して解凍できない
		response=HttpResponse(memory_file.getvalue(),content_type='application/zip')
		# videoidを,で区切ってファイル名にする
		# response['Content-Disposition']=f'attachment; filename="{",".join(request.POST.getlist("videoids"))}.zip"'
		response['Content-Disposition']=f'attachment; filename="test.zip"'

		return response
		# return redirect('app_urls:v1')

	# ajaxに返すjsonレスポンス
	# json_resp={'req_data':req_data,
	# 					 }
	#
	# return JsonResponse(json_resp)